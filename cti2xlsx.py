import os
import tkinter as tk
from tkinter.filedialog import askopenfilenames
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN


def main():
    # Ask for the working files, prn files are optional
    root = tk.Tk()
    root.withdraw()
    print("Select the cti file(s)")
    ctis = askopenfilenames(
        title="Select the cti file(s)", filetypes=[("cti files", ".cti")]
    )
    do_prn = input("Would you like to analyze prn files as well? (Y/N)\n")
    if do_prn == "Y" or do_prn == "y":
        print("Select the prn file(s)")
        prns = askopenfilenames(
            title="Select the prn file(s)", filetypes=[("prn files", ".prn")]
        )

    gave_prn_error = 0
    for x, cti_file in enumerate(ctis, 1):
        print(f"\nWorking on cti file {x} out of {len(ctis)}.")
        cti_name = os.path.basename(cti_file)[:-4]
        # If user wants to analyse prn files as well (do_prn == 1) ¨
        # make sure one with matching name to cti is present
        # Else just act as if do_prn == 0 for this loop
        # (will be managed by found_match)
        found_match = 0
        if do_prn == "Y" or do_prn == "y":
            for p in prns:
                prn_name = os.path.basename(p)[:-8]
                if cti_name == prn_name:
                    print("Found matching prn file")
                    prn_file = p
                    found_match = 1
                    break
            if found_match == 0:
                print(
                    f"Didn't find matching prn file for {cti_name}. Second excel sheet will not be created."
                )
                if gave_prn_error == 0:
                    print(
                        "The names must be EXACTLY the same with prn name having additional '-prn' in the filename."
                    )
                    print(
                        "For example 'abc123.cti' and 'abc123-prn.prn' would get matched."
                    )
                    gave_prn_error = 1

        # Read the cti file as lines
        with open(cti_file) as f:
            lines = f.read().splitlines()
        # Find the indices of beggining and ending lines for each data collumn
        fre_beg = lines.index("VAR_LIST_BEGIN") + 1
        fre_end = lines.index("VAR_LIST_END")
        s11_beg = lines.index("BEGIN", fre_end) + 1
        s11_end = lines.index("END", s11_beg)
        s21_beg = lines.index("BEGIN", s11_end) + 1
        s21_end = lines.index("END", s21_beg)
        s12_beg = lines.index("BEGIN", s21_end) + 1
        s12_end = lines.index("END", s12_beg)
        s22_beg = lines.index("BEGIN", s12_end) + 1
        s22_end = lines.index("END", s22_beg)
        # Put the found data into lists
        # The sXX data have two values on each line, we only care for the first one
        fre_list = lines[fre_beg:fre_end]
        s11_list = [i.split(",")[0] for i in lines[s11_beg:s11_end]]
        s21_list = [i.split(",")[0] for i in lines[s21_beg:s21_end]]
        s12_list = [i.split(",")[0] for i in lines[s12_beg:s12_end]]
        s22_list = [i.split(",")[0] for i in lines[s22_beg:s22_end]]

        # Read the prn file as lines, if it doesn't exist, ignore this whole step
        if found_match == 1:
            with open(prn_file) as f:
                lines = f.read().splitlines()
            # There might be a different number (n) of collumns with info
            # So make a list of lists (with data for each col) of the length n
            num_of_vals = len(lines[3].split())
            prn_data = []
            for n in range(0, num_of_vals):
                temp_list = [i.split()[n] for i in lines[3:]]
                prn_data.append(temp_list)
            for i in range(0, len(prn_data)):
                for j in range(1, len(prn_data[0])):
                    prn_data[i][j] = float(prn_data[i][j])
            for i in range(1, len(prn_data[0])):
                prn_data[0][i] = int(prn_data[0][i])

        # Make an excel file and write the static info
        wb = Workbook()
        ws_cti = wb.active
        ws_cti.title = "cti data"
        ws_cti["A1"].value = cti_name
        ws_cti["C1"].value = "Odrazí"
        ws_cti["D1"].value = "Na druhou"
        ws_cti["F1"].value = "Projde"
        ws_cti["G1"].value = "Na druhou"
        ws_cti["H1"].value = "Absorbuje"
        ws_cti["I1"].value = "Na druhou"
        ws_cti["J1"].value = "Kontrola"
        ws_cti["O1"].value = "SE Total"
        ws_cti["P1"].value = "New Power"
        ws_cti["R1"].value = "SE Total"
        ws_cti["B2"].value = "S11"
        ws_cti["C2"].value = "R (%)"
        ws_cti["D2"].value = "R2 (%)"
        ws_cti["E2"].value = "S21"
        ws_cti["F2"].value = "T (%)"
        ws_cti["G2"].value = "TR (%)"
        ws_cti["H2"].value = "A (%)"
        ws_cti["I2"].value = "A2 (%)"
        ws_cti["J2"].value = "R2 + T2 + A2"
        ws_cti["K2"].value = "S12"
        ws_cti["L2"].value = "S22"
        ws_cti["M2"].value = "SER"
        ws_cti["N2"].value = "SEA"
        ws_cti["O2"].value = "SER + SEA"
        ws_cti["P2"].value = "SER"
        ws_cti["Q2"].value = "SEA"
        ws_cti["R2"].value = "SER + SEA"

        # Make the cells wider
        for i in range(1, 19):
            ws_cti.column_dimensions[get_column_letter(i)].width = 13
            ws_cti[get_column_letter(i) + "2"].border = Border(
                bottom=Side(border_style=BORDER_THIN, color="00000000")
            )

        # Colour the cells
        col_yel = "FFE699"
        col_grn = "C6E0B4"
        col_blu = "BDD7EE"
        col_red = "F8CBAD"
        col_pin = "FF9999"
        col_alr = "FFFF99"

        ws_cti["A1"].fill = PatternFill(fgColor=col_yel, fill_type="solid")
        ws_cti["C1"].fill = PatternFill(fgColor=col_grn, fill_type="solid")
        ws_cti["D1"].fill = PatternFill(fgColor=col_grn, fill_type="solid")
        ws_cti["C2"].fill = PatternFill(fgColor=col_grn, fill_type="solid")
        ws_cti["D2"].fill = PatternFill(fgColor=col_grn, fill_type="solid")
        ws_cti["F1"].fill = PatternFill(fgColor=col_blu, fill_type="solid")
        ws_cti["G1"].fill = PatternFill(fgColor=col_blu, fill_type="solid")
        ws_cti["F2"].fill = PatternFill(fgColor=col_blu, fill_type="solid")
        ws_cti["G2"].fill = PatternFill(fgColor=col_blu, fill_type="solid")
        ws_cti["H1"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["I1"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["J1"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["H2"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["I2"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["J2"].fill = PatternFill(fgColor=col_red, fill_type="solid")
        ws_cti["P1"].fill = PatternFill(fgColor=col_pin, fill_type="solid")
        ws_cti["Q1"].fill = PatternFill(fgColor=col_pin, fill_type="solid")
        ws_cti["R1"].fill = PatternFill(fgColor=col_pin, fill_type="solid")
        ws_cti["P2"].fill = PatternFill(fgColor=col_pin, fill_type="solid")
        ws_cti["Q2"].fill = PatternFill(fgColor=col_pin, fill_type="solid")
        ws_cti["R2"].fill = PatternFill(fgColor=col_pin, fill_type="solid")

        # Write cti values
        last_ghz = str(int(fre_list[0]))[:-9]
        for i in range(0, len(fre_list)):
            ws_cti["A" + str(i + 3)].value = float(fre_list[i])
            ws_cti["B" + str(i + 3)].value = float(s11_list[i])
            ws_cti["C" + str(i + 3)].value = f"=POWER(10,(B{i + 3}/20))"
            ws_cti["D" + str(i + 3)].value = f"=POWER(C{i + 3},2)*(100)"
            ws_cti["E" + str(i + 3)].value = float(s21_list[i])
            ws_cti["F" + str(i + 3)].value = f"=POWER(10,(E{i + 3}/20))"
            ws_cti["G" + str(i + 3)].value = f"=POWER(F{i + 3},2)*(100)"
            ws_cti[
                "H" + str(i + 3)
            ].value = f"=SQRT(1-((POWER(C{i + 3},2))+(POWER(F{i + 3},2))))"
            ws_cti["I" + str(i + 3)].value = f"=POWER(H{i + 3},2)*(100)"
            ws_cti["J" + str(i + 3)].value = f"=(D{i + 3}+G{i + 3}+I{i + 3})"
            ws_cti["K" + str(i + 3)].value = float(s12_list[i])
            ws_cti["L" + str(i + 3)].value = float(s22_list[i])
            ws_cti["M" + str(i + 3)].value = f"=ABS(10*LOG(1/ABS(1-B{i + 3}^2)))"
            ws_cti[
                "N" + str(i + 3)
            ].value = f"=ABS(10*LOG(ABS((1-B{i + 3}^2)/K{i + 3}^2)))"
            ws_cti["O" + str(i + 3)].value = f"=ABS(M{i + 3}+N{i + 3})"
            ws_cti["P" + str(i + 3)].value = f"=10*LOG(1/(1-10^(B{i + 3}/10)))"
            ws_cti[
                "Q" + str(i + 3)
            ].value = f"=10*LOG((1-10^(B{i + 3}/10))/10^(E{i + 3}/10))"
            ws_cti["R" + str(i + 3)].value = f"=ABS(P{i + 3}+Q{i + 3})"
            # Colour the cells for each collumn that needs it
            ws_cti["C" + str(i + 3)].fill = PatternFill(
                fgColor=col_grn, fill_type="solid"
            )
            ws_cti["D" + str(i + 3)].fill = PatternFill(
                fgColor=col_grn, fill_type="solid"
            )
            ws_cti["F" + str(i + 3)].fill = PatternFill(
                fgColor=col_blu, fill_type="solid"
            )
            ws_cti["G" + str(i + 3)].fill = PatternFill(
                fgColor=col_blu, fill_type="solid"
            )
            ws_cti["H" + str(i + 3)].fill = PatternFill(
                fgColor=col_red, fill_type="solid"
            )
            ws_cti["I" + str(i + 3)].fill = PatternFill(
                fgColor=col_red, fill_type="solid"
            )
            ws_cti["J" + str(i + 3)].fill = PatternFill(
                fgColor=col_red, fill_type="solid"
            )
            ws_cti["P" + str(i + 3)].fill = PatternFill(
                fgColor=col_pin, fill_type="solid"
            )
            ws_cti["q" + str(i + 3)].fill = PatternFill(
                fgColor=col_pin, fill_type="solid"
            )
            ws_cti["r" + str(i + 3)].fill = PatternFill(
                fgColor=col_pin, fill_type="solid"
            )
            # Yellow the line when frequency ghz changes
            ghz = str(int(fre_list[i]))[:-9]
            if ghz != last_ghz:
                for j in range(1, 19):
                    ws_cti[get_column_letter(j) + str(i + 3)].fill = PatternFill(
                        fgColor=col_alr, fill_type="solid"
                    )
            last_ghz = ghz

        # Open the prn workseet if prn file was foudn and write the prn values
        # Also do all the cell eiditng like in cti worksheet
        if found_match == 1:
            ws_prn = wb.create_sheet("prn data")
            ws_prn["A1"].value = cti_name
            ws_prn["A1"].fill = PatternFill(fgColor=col_yel, fill_type="solid")
            for i in range(1, 7):
                ws_prn.column_dimensions[get_column_letter(i)].width = 13
                ws_prn[get_column_letter(i) + "2"].border = Border(
                    bottom=Side(border_style=BORDER_THIN, color="00000000")
                )
            for i in range(0, len(prn_data)):
                for j in range(0, len(prn_data[0])):
                    ws_prn[get_column_letter(i + 2) + str(j + 2)].value = prn_data[i][j]

        # Save and close the excel file
        wb.save(cti_file[:-4] + ".xlsx")
        wb.close()

    print("\nAll done\nPress any key to exit")
    input()


if __name__ == "__main__":
    main()
